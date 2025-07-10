import pandas as pd
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from datetime import datetime, timedelta
import re
import os
import glob
from tkinter import *
from tkinter import ttk, filedialog, messagebox
import threading
import subprocess
import sys

# 导入中文大写数字转换函数
def num_to_chinese(num):
    """
    将数字转换为中文大写金额
    """
    # 特殊情况处理
    if num == 0:
        return '零圆整'
    
    num = float(num)
    integer_part = int(num)
    decimal_part = int(round((num - integer_part) * 100))
    
    chinese_nums = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    position_units = ['', '拾', '佰', '仟']  # 个位不添加单位，后面单独处理
    section_units = ['', '万', '亿', '兆', '京', '垓']
    
    # 处理整数部分
    chinese_str = ''
    
    # 特殊情况：整数部分为0
    if integer_part == 0:
        chinese_str = '零圆'
    else:
        # 将整数部分转换为字符串
        str_integer = str(integer_part)
        
        # 按4位分段，从低位到高位
        sections = []
        for i in range(0, len(str_integer), 4):
            start = max(0, len(str_integer) - i - 4)
            end = len(str_integer) - i
            sections.append(str_integer[start:end])
        
        # 处理每个分段
        for section_index, section in enumerate(sections):
            section_chinese = ''
            has_value = False  # 标记这一段是否有非零值
            
            # 处理每一段内的数字，从高位到低位
            for i, digit in enumerate(section):
                position = len(section) - i - 1  # 位置（个、十、百、千）
                digit_int = int(digit)
                
                if digit_int != 0:
                    # 添加数字和单位
                    section_chinese += chinese_nums[digit_int] + position_units[position]
                    has_value = True
                elif has_value:  # 如果之前有非零值，且当前是零
                    # 避免多个连续的零
                    if not section_chinese.endswith('零'):
                        section_chinese += '零'
            
            # 处理末尾的零
            if section_chinese.endswith('零'):
                section_chinese = section_chinese[:-1]
            
            # 如果这一段有内容，添加万、亿等单位
            if section_chinese != '':
                if section_index < len(section_units):
                    section_chinese += section_units[section_index]
                chinese_str = section_chinese + chinese_str
        
        # 在整数部分的最后添加"圆"字（即个位数后面）
        chinese_str += '圆'
    
    # 处理小数部分
    if decimal_part > 0:
        jiao = decimal_part // 10
        fen = decimal_part % 10
        
        if jiao > 0:
            chinese_str += chinese_nums[jiao] + '角'
        if fen > 0:
            chinese_str += chinese_nums[fen] + '分'
    else:
        # 只有在没有小数部分时才添加"整"字
        chinese_str += '整'
    
    # 确保结果不为空
    if not chinese_str:
        chinese_str = '零圆整'
    
    return chinese_str

# 忽略来自openpyxl.styles.stylesheet的UserWarning
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl.styles.stylesheet')

class ProductClassificationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("供应商对帐确认函")
        
        # 设置窗口大小并居中
        self.set_window_geometry(600, 650)
        
        # 使窗口前台显示
        self.bring_to_front()
        
        # 检查时间验证
        if not self.check_expiration():
            messagebox.showerror("错误", "程序已过期，请联系开发者")
            self.root.destroy()
            return
        
        # 创建主框架
        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.pack(fill=BOTH, expand=True)
        
        # 创建控制面板
        self.create_control_panel()
        
        # 创建日志显示区域
        self.create_log_area()
        
        # 初始化状态
        self.processing = False
        
        # 创建开发者信息标签
        self.create_developer_label()
    
    def set_window_geometry(self, width, height):
        """设置窗口大小并居中"""
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        self.root.geometry(f"{width}x{height}+{x}+{y}")
    
    def check_expiration(self):
        """检查时间是否到期"""
        current_date = datetime.now()
        expiration_date = datetime(2025, 12, 31)  # 2025年底到期
        
        return current_date <= expiration_date
    
    def create_control_panel(self):
        control_frame = ttk.LabelFrame(self.main_frame, text="生成标记及确认函", padding="10")
        control_frame.pack(fill=X, pady=5)
        
        # 选择模式框架
        mode_frame = ttk.Frame(control_frame)
        mode_frame.pack(fill=X, pady=5)
        
        ttk.Label(mode_frame, text="选择模式:").pack(side=LEFT, padx=(0, 10))
        
        # 选择模式单选按钮组
        self.mode_var = StringVar(value="multi_files")
        
        multi_radio = ttk.Radiobutton(mode_frame, text="文件", variable=self.mode_var, 
                                     value="multi_files", command=self.update_file_selection_ui)
        multi_radio.pack(side=LEFT, padx=5)
        
        folder_radio = ttk.Radiobutton(mode_frame, text="文件夹", variable=self.mode_var, 
                                      value="folder", command=self.update_file_selection_ui)
        folder_radio.pack(side=LEFT, padx=5)
        
        # 添加在原文件上操作的选项
        option_frame = ttk.Frame(control_frame)
        option_frame.pack(fill=X, pady=5)
        
        self.edit_in_place_var = BooleanVar(value=False)
        edit_in_place_check = ttk.Checkbutton(option_frame, text="直接在原文件上操作", 
                                             variable=self.edit_in_place_var)
        edit_in_place_check.pack(side=LEFT, padx=5)
        
        # 文件选择框架
        self.file_selection_frame = ttk.Frame(control_frame)
        self.file_selection_frame.pack(fill=X, pady=5)
        
        # 初始化文件选择UI
        self.update_file_selection_ui()
        
        # 处理按钮
        self.process_btn = ttk.Button(control_frame, text="开始处理", command=self.start_processing)
        self.process_btn.pack(pady=10)
        
        # 进度条
        self.progress = ttk.Progressbar(control_frame, orient=HORIZONTAL, mode='determinate')
        self.progress.pack(fill=X, pady=5)
    
    def create_log_area(self):
        log_frame = ttk.LabelFrame(self.main_frame, text="处理日志", padding="10")
        log_frame.pack(fill=BOTH, expand=True)
        
        self.log_text = Text(log_frame, wrap=WORD, state=DISABLED)
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=RIGHT, fill=Y)
        self.log_text.pack(fill=BOTH, expand=True)
    
    def update_file_selection_ui(self, *args):
        """根据选择的模式更新文件选择UI"""
        # 清空当前框架中的所有控件
        for widget in self.file_selection_frame.winfo_children():
            widget.destroy()
        
        mode = self.mode_var.get()
        
        if mode == "multi_files":
            # 文件选择UI
            ttk.Label(self.file_selection_frame, text="选择Excel文件:").pack(side=LEFT)
            self.input_files_var = StringVar()
            ttk.Entry(self.file_selection_frame, textvariable=self.input_files_var, width=40).pack(side=LEFT, padx=5)
            ttk.Button(self.file_selection_frame, text="浏览...", command=self.select_input_files).pack(side=LEFT)
        
        elif mode == "folder":
            # 文件夹选择UI
            ttk.Label(self.file_selection_frame, text="选择文件夹:").pack(side=LEFT)
            self.input_folder_var = StringVar()
            ttk.Entry(self.file_selection_frame, textvariable=self.input_folder_var, width=40).pack(side=LEFT, padx=5)
            ttk.Button(self.file_selection_frame, text="浏览...", command=self.select_input_folder).pack(side=LEFT)
    

    
    def select_input_files(self):
        """选择多个文件"""
        filetypes = [("Excel files", "*.xlsx *.xls")]
        file_paths = filedialog.askopenfilenames(filetypes=filetypes)
        if file_paths:
            self.input_files_var.set(";;".join(file_paths))  # 使用双分号作为分隔符，避免路径中的单分号冲突
    
    def select_input_folder(self):
        """选择文件夹"""
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.input_folder_var.set(folder_path)
    
    def log_message(self, message):
        """添加消息到日志区域"""
        self.log_text.config(state=NORMAL)
        # 配置警告和错误标签为红色
        self.log_text.tag_config("warning", foreground="red")
        
        # 检查消息是否包含警告、失败、错误或其他问题关键词
        error_keywords = ["警告", "失败", "错误", "出错", "无法", "异常", "Exception", "[失败]", "不存在"]
        is_error = False
        
        # 检查消息中是否包含任何错误关键词
        for keyword in error_keywords:
            if keyword in message:
                is_error = True
                break
        
        if is_error:
            self.log_text.insert(END, message + "\n", "warning")
        else:
            self.log_text.insert(END, message + "\n")
        self.log_text.see(END)
        self.log_text.config(state=DISABLED)
    
    def start_processing(self):
        if self.processing:
            return
        
        mode = self.mode_var.get()
        files_to_process = []
        
        # 根据不同模式获取要处理的文件
        if mode == "multi_files":
            input_files = self.input_files_var.get()
            if not input_files:
                messagebox.showwarning("警告", "请先选择Excel文件")
                return
            files_to_process = input_files.split(";;")  # 使用双分号分隔符
            
        elif mode == "folder":
            input_folder = self.input_folder_var.get()
            if not input_folder:
                messagebox.showwarning("警告", "请先选择文件夹")
                return
                
            # 查找所有Excel文件
            excel_files = glob.glob(os.path.join(input_folder, "*.xlsx")) + glob.glob(os.path.join(input_folder, "*.xls"))
            files_to_process.extend(excel_files)
            
            if not files_to_process:
                messagebox.showwarning("警告", f"在文件夹 '{input_folder}' 中没有找到Excel文件")
                return
        
        # 去除重复文件
        files_to_process = list(set(files_to_process))
        
        self.processing = True
        self.process_btn.config(state=DISABLED)
        self.log_text.config(state=NORMAL)
        self.log_text.delete(1.0, END)
        self.log_text.config(state=DISABLED)
        self.progress['value'] = 0
        
        # 使用线程处理，避免界面卡顿
        threading.Thread(target=self.process_multiple_files, args=(files_to_process,), daemon=True).start()
    
    def process_multiple_files(self, file_paths):
        """处理多个文件"""
        try:
            total_files = len(file_paths)
            self.log_message(f"共找到 {total_files} 个文件需要处理")
            
            # 初始化统计信息
            successful_files = 0
            failed_files = 0
            
            # 处理每个文件
            for i, file_path in enumerate(file_paths):
                # 更新总体进度
                overall_progress = int((i / total_files) * 100)
                self.progress['value'] = overall_progress
                self.root.update_idletasks()
                
                # 处理单个文件
                self.log_message(f"\n[{i+1}/{total_files}] 开始处理文件: {os.path.basename(file_path)}")
                
                # 调用处理单个文件的方法
                success = self.process_file(file_path, is_batch=True)
                
                if success:
                    successful_files += 1
                    self.log_message(f"[成功] 文件 {os.path.basename(file_path)} 处理完成")
                else:
                    failed_files += 1
                    self.log_message(f"[失败] 文件 {os.path.basename(file_path)} 处理失败")
            
            # 更新进度条到100%
            self.progress['value'] = 100
            
            # 显示处理汇总信息
            self.log_message(f"\n处理完成汇总:")
            self.log_message(f"总文件数: {total_files}")
            self.log_message(f"成功处理: {successful_files}")
            self.log_message(f"处理失败: {failed_files}")
            
            if successful_files > 0:
                # 获取输出目录（假设所有文件都在同一个目录）
                output_dir = os.path.dirname(file_paths[0])
                
                message = f"共处理 {total_files} 个文件，成功 {successful_files} 个，失败 {failed_files} 个。"
                if self.edit_in_place_var.get():
                    message += "\n\n已直接在原文件上操作。"
                else:
                    message += "\n\n已保存为新文件。"
                
                if messagebox.askyesno("处理完成", f"{message}\n\n是否打开输出文件夹？"):
                    try:
                        if sys.platform == "win32":
                            os.startfile(output_dir)
                        elif sys.platform == "darwin":  # macOS
                            subprocess.call(["open", output_dir])
                        else:  # Linux
                            subprocess.call(["xdg-open", output_dir])
                    except Exception as e:
                        self.log_message(f"无法打开文件夹: {str(e)}")
                        messagebox.showerror("错误", f"无法打开文件夹:\n{str(e)}")
            else:
                messagebox.showwarning("处理失败", "所有文件处理失败，请检查文件格式是否正确")
                
        except Exception as e:
            self.log_message(f"批量处理文件时出错: {str(e)}")
            messagebox.showerror("错误", f"批量处理文件时出错:\n{str(e)}")
        finally:
            self.processing = False
            self.process_btn.config(state=NORMAL)
    
    def process_file(self, file_path, is_batch=False):
        """处理单个文件，返回是否成功。当is_batch=True时，作为批处理模式的一部分运行，不显示单独的消息框"""
        try:
            if not is_batch:
                self.log_message(f"开始处理文件: {os.path.basename(file_path)}")
            
            # 检查文件是否存在
            if not os.path.exists(file_path):
                self.log_message("警告：文件不存在")
                if not is_batch:
                    messagebox.showerror("错误", "选择的文件不存在")
                    self.processing = False
                    self.process_btn.config(state=NORMAL)
                return False
            
            # 读取Excel文件
            self.log_message("读取Excel文件...")
            try:
                # 表头在第6行，所以跳过前5行
                df = pd.read_excel(file_path, header=5)
                self.log_message(f"成功读取文件，共 {len(df)} 行数据")
            except Exception as e:
                self.log_message(f"警告：读取Excel文件失败: {str(e)}")
                if not is_batch:
                    messagebox.showerror("错误", f"无法读取Excel文件:\n{str(e)}")
                    self.processing = False
                    self.process_btn.config(state=NORMAL)
                return False
            
            # 检查是否存在M列（Excel中的第13列）
            if len(df.columns) < 13:  # 假设M列是第13列（索引为12）
                self.log_message("警告：文件中没有足够的列，无法找到M列")
                if not is_batch:
                    self.processing = False
                    self.process_btn.config(state=NORMAL)
                return False
            
            # 获取M列的列名和数据
            m_column_name = df.columns[12]  # 索引为12的列（M列）
            self.log_message(f"找到M列: {m_column_name}")
            
            # 添加新列用于存储分类结果（在M列旁边）
            classification_column = "品类标记"
            df.insert(13, classification_column, "")  # 在M列后插入新列，默认为空
            
            # 进行分类标记
            total_rows = len(df)
            for i, row in df.iterrows():
                # 更新进度条
                progress_value = int((i + 1) / total_rows * 100)
                self.progress['value'] = progress_value
                self.root.update_idletasks()  # 强制更新UI
                
                # 获取M列内容
                m_value = str(row[m_column_name]) if pd.notna(row[m_column_name]) else ""
                
                # 如果M列内容为空，则不进行标记
                if not m_value:
                    continue
                
                # 应用分类规则
                # 1. 干货：准确查找M列内容有"鱼虾蟹干及瑶柱干"，"海参鲍鱼鱼翅干及肚干"，"其他水产干货"。"燕窝"将被标记为干货。
                if any(keyword in m_value for keyword in ["鱼虾蟹干及瑶柱干", "海参鲍鱼鱼翅干及肚干", "其他水产干货"]) or "燕窝" in m_value:
                    df.at[i, classification_column] = "干货"
                # 2. 海鲜：M列内容包含"活鲜"2个字，即被标记为海鲜
                elif "活鲜" in m_value:
                    df.at[i, classification_column] = "海鲜"
                # 3. 酒类：M列内容包含"酒"1个字，将被标记为酒类
                elif "酒" in m_value:
                    df.at[i, classification_column] = "酒类"
                # 4. 饮料：M列内容包含"饮料"2个字，即被标记为饮料
                elif "饮料" in m_value:
                    df.at[i, classification_column] = "饮料"
                # 5. 水：M列内容只有"水"这个字，即被标记为水
                elif m_value == "水":
                    df.at[i, classification_column] = "水"
                # 6. 其他：所有未被以上标记的商品，将被标记为其他。
                else:
                    df.at[i, classification_column] = "其他"
            
            # 根据用户选择决定是保存到新文件还是直接修改原文件
            if self.edit_in_place_var.get():
                output_file = file_path
                self.log_message("将直接在原文件上操作...")
            else:
                output_dir = os.path.dirname(file_path)
                file_name, file_ext = os.path.splitext(os.path.basename(file_path))
                output_file = os.path.join(output_dir, f"{file_name}_分类{file_ext}")
                self.log_message("正在保存到新文件...")
            
            try:
                # 尝试使用openpyxl保存，保留原始格式
                # 先读取原始文件以保留格式
                try:
                    wb = load_workbook(file_path)
                    ws = wb.active
                    
                    # 尝试读取Statement Sheet中的L7单元格数据（供应商名称）
                    supplier_name = ""
                    try:
                        # 检查是否存在名为"Statement Sheet"的工作表
                        if "Statement Sheet" in wb.sheetnames:
                            statement_sheet = wb["Statement Sheet"]
                            supplier_name = statement_sheet.cell(row=7, column=12).value  # L列是第12列
                            if supplier_name:
                                self.log_message(f"从Statement Sheet的L7单元格读取到供应商名称: {supplier_name}")
                            else:
                                self.log_message("Statement Sheet的L7单元格没有数据")
                        else:
                            # 如果没有Statement Sheet，尝试从第一个工作表的L7单元格读取
                            supplier_name = ws.cell(row=7, column=12).value  # L列是第12列
                            if supplier_name:
                                self.log_message(f"从第一个工作表的L7单元格读取到供应商名称: {supplier_name}")
                            else:
                                self.log_message("第一个工作表的L7单元格没有数据")
                    except Exception as e:
                        self.log_message(f"读取供应商名称时出错: {str(e)}")
                        supplier_name = ""
                    
                    # 添加新列标题
                    header_row = 6  # 表头在第6行
                    ws.cell(row=header_row, column=14, value=classification_column)
                    
                    # 添加分类结果
                    for i, row in df.iterrows():
                        ws.cell(row=i+7, column=14, value=row[classification_column])  # +7是因为Excel行从1开始，且表头在第6行
                    
                    # 创建汇总sheet
                    if "汇总" not in wb.sheetnames:
                        summary_sheet = wb.create_sheet(title="汇总")
                    else:
                        summary_sheet = wb["汇总"]
                    
                    # 设置页面边距和页眉页脚（单位：厘米）
                    summary_sheet.page_margins = PageMargins(top=0.5/2.54, left=1.5/2.54, right=0.5/2.54, bottom=0.5/2.54, header=0, footer=0)
                    summary_sheet.page_setup.horizontalCentered = True
                    
                    # 设置汇总sheet的标题
                    summary_sheet.cell(row=1, column=1, value="供应商对账确认函")
                    summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=16)
                    summary_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    # 合并标题单元格
                    summary_sheet.merge_cells('A1:F1')
                    
                    # 读取config.txt文件获取酒店信息
                    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.txt")
                    hotel_name = ""
                    hotel_address = ""
                    contact_person = ""
                    email_address = ""
                    
                    if os.path.exists(config_path):
                        try:
                            with open(config_path, 'r', encoding='utf-8') as f:
                                for line in f:
                                    line = line.strip()
                                    if line.startswith("B2:"):
                                        hotel_name = line.replace("B2:", "").strip()
                                    elif line.startswith("D2:"):
                                        hotel_address = line.replace("D2:", "").strip()
                                    elif line.startswith("E2:"):
                                        contact_person = line.replace("E2:", "").strip()
                                    elif line.startswith("B43:"):
                                        email_address = line.replace("B43:", "").strip()
                            self.log_message(f"已从config.txt读取酒店信息")
                        except Exception as e:
                            self.log_message(f"读取config.txt失败: {str(e)}")
                    
                    # 在第二行开始插入文字
                    summary_sheet.cell(row=2, column=1, value="由酒店（酒店全称）：")
                    summary_sheet.cell(row=2, column=2, value=hotel_name)
                    summary_sheet.cell(row=3, column=1, value="地址：")
                    summary_sheet.cell(row=3, column=2, value=hotel_address)
                    summary_sheet.cell(row=4, column=1, value="财务部联系人：")
                    summary_sheet.cell(row=4, column=2, value=contact_person)
                    summary_sheet.cell(row=5, column=1, value="致供应商（供应商全称）：")
                    # 将从Statement Sheet读取的供应商名称写入B5单元格
                    summary_sheet.cell(row=5, column=2, value=supplier_name)
                    summary_sheet.cell(row=6, column=1, value="税务登记号码：")
                    summary_sheet.cell(row=7, column=1, value="对账联系人：")
                    summary_sheet.cell(row=8, column=1, value="经酒店与供应商共同核对，确认产生如下交易货款：")
                    summary_sheet.cell(row=9, column=1, value="➢ 含税总金额人民币大写：")
                    summary_sheet.cell(row=10, column=1, value="➢ 不含税金额：")
                    summary_sheet.cell(row=11, column=1, value="➢ 增值税税款：")
                    summary_sheet.cell(row=12, column=1, value="货款所属期间：")
                    summary_sheet.cell(row=13, column=1, value="明细对账信息如下：")
                    
                    # 合并第2-7行的B-D列
                    for row in range(2, 8):
                        summary_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                        # 移除背景色
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.fill = PatternFill(fill_type=None)
                    
                    # 合并第9-13行的B-D列
                    for row in range(9, 14):
                        summary_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                        # 移除背景色
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.fill = PatternFill(fill_type=None)
                    
                    # 创建新的表格结构，与图片中的表格结构一致
                    # 表头第一行
                    summary_sheet.cell(row=14, column=1, value="")
                    summary_sheet.merge_cells(start_row=14, start_column=1, end_row=15, end_column=1)
                    
                    summary_sheet.cell(row=14, column=2, value="员餐")
                    summary_sheet.merge_cells(start_row=14, start_column=2, end_row=14, end_column=3)
                    
                    summary_sheet.cell(row=14, column=4, value="其他餐饮点 - 非员餐")
                    summary_sheet.merge_cells(start_row=14, start_column=4, end_row=14, end_column=5)
                    
                    summary_sheet.cell(row=14, column=6, value="当月总应付账款金额")
                    summary_sheet.merge_cells(start_row=14, start_column=6, end_row=15, end_column=6)
                    
                    # 表头第二行
                    summary_sheet.cell(row=15, column=2, value="不含税金额")
                    summary_sheet.cell(row=15, column=3, value="税费")
                    summary_sheet.cell(row=15, column=4, value="不含税金额")
                    summary_sheet.cell(row=15, column=5, value="税费")
                    
                    # 设置品类列标题
                    summary_sheet.cell(row=14, column=1, value="品类")

                    
                    # 设置表头样式
                    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    for row in range(14, 16):  # 修改为只包含第14-15行
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.fill = header_fill
                            
                            # 添加边框
                            from openpyxl.styles import Border, Side
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            cell.border = thin_border
                    
                    # 按用户要求的顺序显示所有分类
                    ordered_categories = ["干货", "海鲜", "酒类", "饮料", "水", "其他"]
                    row_idx = 16  # 从第16行开始填充数据（表头占据14-15行）
                    
                    # 定义员工餐厅和其他餐厅（营业点）
                    employee_restaurants = ["员工餐厅", "员工食堂"]
                    
                    # 初始化总计变量
                    total_employee_untaxed = 0
                    total_employee_tax = 0
                    total_other_untaxed = 0
                    total_other_tax = 0
                    
                    # 直接填充各分类数据到新表格结构
                    for category in ordered_categories:
                        # 筛选该分类的员工餐厅数据
                        employee_df = df[(df[classification_column] == category) & 
                                         (df["部门"].isin(employee_restaurants))]
                        
                        # 计算员工餐厅未税金额和税额
                        employee_untaxed = employee_df["小计金额(结算)"].sum() if not employee_df.empty else 0
                        employee_tax = employee_df["税额(结算)"].sum() if not employee_df.empty else 0
                        
                        # 更新员工餐厅总计
                        total_employee_untaxed += employee_untaxed
                        total_employee_tax += employee_tax
                        
                        # 筛选该分类的其他餐厅（非员餐）数据
                        other_df = df[(df[classification_column] == category) & 
                                      (~df["部门"].isin(employee_restaurants))]
                        
                        # 计算其他餐厅未税金额和税额
                        other_untaxed = other_df["小计金额(结算)"].sum() if not other_df.empty else 0
                        other_tax = other_df["税额(结算)"].sum() if not other_df.empty else 0
                        
                        # 更新其他餐厅总计
                        total_other_untaxed += other_untaxed
                        total_other_tax += other_tax
                        
                        # 计算当月总应付账款金额
                        total_row_amount = employee_untaxed + employee_tax + other_untaxed + other_tax
                        
                        # 写入汇总数据
                        summary_sheet.cell(row=row_idx, column=1, value=category)
                        summary_sheet.cell(row=row_idx, column=2, value="-" if employee_untaxed == 0 else employee_untaxed)
                        summary_sheet.cell(row=row_idx, column=3, value="-" if employee_tax == 0 else employee_tax)
                        summary_sheet.cell(row=row_idx, column=4, value="-" if other_untaxed == 0 else other_untaxed)
                        summary_sheet.cell(row=row_idx, column=5, value="-" if other_tax == 0 else other_tax)
                        summary_sheet.cell(row=row_idx, column=6, value="-" if total_row_amount == 0 else total_row_amount)
                        
                        # 设置单元格样式
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row_idx, column=col)
                            if col > 1:  # 数字列设置数字格式
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            else:  # 品类列左对齐
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            
                            # 添加边框
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            cell.border = thin_border
                        
                        row_idx += 1
                        
                    # 添加总计行
                    summary_sheet.cell(row=row_idx, column=1, value="合计")
                    summary_sheet.cell(row=row_idx, column=2, value="-" if total_employee_untaxed == 0 else total_employee_untaxed)
                    summary_sheet.cell(row=row_idx, column=3, value="-" if total_employee_tax == 0 else total_employee_tax)
                    summary_sheet.cell(row=row_idx, column=4, value="-" if total_other_untaxed == 0 else total_other_untaxed)
                    summary_sheet.cell(row=row_idx, column=5, value="-" if total_other_tax == 0 else total_other_tax)
                    
                    # 计算总金额
                    total_amount = total_employee_untaxed + total_employee_tax + total_other_untaxed + total_other_tax
                    summary_sheet.cell(row=row_idx, column=6, value="-" if total_amount == 0 else total_amount)
                    
                    # 设置总计行样式
                    total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                    for col in range(1, 7):
                        cell = summary_sheet.cell(row=row_idx, column=col)
                        cell.font = Font(bold=True, size=12)
                        cell.fill = total_fill
                        
                        # 设置底部双边框
                        from openpyxl.styles import Border, Side
                        double_bottom_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='double')
                        )
                        cell.border = double_bottom_border
                        
                        if col == 1:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            cell.number_format = '#,##0.00'
                    
                    # 读取总计行的第6列（总金额）并转换为中文大写写入B9单元格
                    try:
                        # total_amount已在前面计算
                        if total_amount is not None:
                            # 转换为中文大写（函数内部已添加"圆"字）
                            chinese_amount = num_to_chinese(total_amount)
                            # 转换为小写
                            lowercase_amount = f"{total_amount:.2f}元"
                            # 写入B9单元格（含税总金额人民币大写）
                            summary_sheet.cell(row=9, column=2, value=f"{chinese_amount}（{lowercase_amount}）")
                            self.log_message(f"已将总金额 {total_amount} 转换为大写 {chinese_amount} 并写入B9单元格")
                        else:
                            self.log_message("总金额为空，无法转换为中文大写")
                    except Exception as e:
                        self.log_message(f"转换总金额为中文大写时出错: {str(e)}")
                        # 如果出错，尝试直接写入原始值
                        try:
                            if total_amount is not None:
                                summary_sheet.cell(row=9, column=2, value=f"{total_amount:.2f}元")
                        except:
                            pass
                    
                    # 读取总计行的数据并写入B10和B11单元格
                    try:
                        # 使用当前总计行的数据
                        total_untaxed = total_employee_untaxed + total_other_untaxed
                        total_tax = total_employee_tax + total_other_tax
                        
                        if total_untaxed is not None:
                            # 写入B10单元格，前面加上"小写"，后面加上"元"
                            summary_sheet.cell(row=10, column=2, value=f"小写{total_untaxed:.2f}元")
                            self.log_message(f"已将未税总金额 {total_untaxed} 写入B10单元格")
                        else:
                            self.log_message("未税总金额为空，无法写入B10单元格")
                            
                        if total_tax is not None:
                            # 写入B11单元格，前面加上"小写"，后面加上"元"
                            summary_sheet.cell(row=11, column=2, value=f"小写{total_tax:.2f}元")
                            self.log_message(f"已将税额总金额 {total_tax} 写入B11单元格")
                        else:
                            self.log_message("税额总金额为空，无法写入B11单元格")
                    except Exception as e:
                        self.log_message(f"读取总计行数据并写入B10和B11单元格时出错: {str(e)}")
                        # 如果出错，记录错误但继续执行
                    
                    # 读取Statement sheet中的A列年月数据并转换格式写入B12单元格
                    try:
                        # 获取年月数据
                        year_month = ""
                        # 检查是否存在名为"Statement Sheet"的工作表
                        if "Statement Sheet" in wb.sheetnames:
                            statement_sheet = wb["Statement Sheet"]
                            # 尝试从A列获取年月数据（通常在A1或其他位置）
                            for row in range(1, 10):  # 检查前10行
                                cell_value = statement_sheet.cell(row=row, column=1).value
                                if cell_value and isinstance(cell_value, str) and re.search(r'\d{4}[-年]\d{1,2}', cell_value):
                                    year_month = cell_value
                                    break
                        
                        # 如果没有找到年月数据，尝试从文件名获取
                        if not year_month:
                            file_name = os.path.basename(file_path)
                            match = re.match(r'(\d{4}-\d{2})_(.+?)(_分类)?\.xlsx', file_name)
                            if match:
                                year_month = match.group(1)
                        
                        # 如果仍然没有找到年月数据，使用当前年月
                        if not year_month:
                            now = datetime.now()
                            year_month = now.strftime('%Y-%m')
                        
                        # 解析年月数据
                        if '-' in year_month:
                            year, month = year_month.split('-')
                        elif '年' in year_month:
                            match = re.search(r'(\d{4})年(\d{1,2})', year_month)
                            if match:
                                year, month = match.group(1), match.group(2)
                            else:
                                raise ValueError(f"无法解析年月格式: {year_month}")
                        else:
                            raise ValueError(f"无法解析年月格式: {year_month}")
                        
                        # 获取月份的最后一天
                        if int(month) == 12:
                            next_month = datetime(int(year) + 1, 1, 1)
                        else:
                            next_month = datetime(int(year), int(month) + 1, 1)
                        
                        last_day = (next_month - timedelta(days=1)).day
                        
                        # 格式化为"2025年6月1日至2025年6月30日"格式
                        formatted_date = f"{year}年{month}月1日至{year}年{month}月{last_day}日"
                        
                        # 写入B12单元格
                        summary_sheet.cell(row=12, column=2, value=formatted_date)
                        self.log_message(f"已将年月数据转换为 {formatted_date} 并写入B12单元格")
                    except Exception as e:
                        self.log_message(f"读取年月数据并转换格式写入B12单元格时出错: {str(e)}")
                        # 如果出错，记录错误但继续执行
                    
                    # 调整列宽
                    summary_sheet.column_dimensions["A"].width = 28
                    summary_sheet.column_dimensions["B"].width = 15
                    summary_sheet.column_dimensions["C"].width = 12
                    summary_sheet.column_dimensions["D"].width = 12
                    summary_sheet.column_dimensions["E"].width = 12
                    summary_sheet.column_dimensions["F"].width = 20
                    # 在A25单元格开始插入备注文字
                    summary_sheet.cell(row=25, column=1, value="备注：")
                    summary_sheet.cell(row=25, column=1).font = Font(bold=True)
                    # 合并A25-F25单元格
                    summary_sheet.merge_cells(start_row=25, start_column=1, end_row=25, end_column=6)
                    
                    # 设置备注文字的样式
                    remark_font = Font(size=11)
                    remark_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # 添加备注内容
                    remarks = [
                        "1. 品类根据供应商实际送货的情况填写，不适用的可留空",
                        "2. 员餐货款的不含税金额，如零税率，酒店需要根据实际收货记录的总金额去换算含税及不含税填写",
                        "3. 本函由双方核对原始收货单据后填写，供应商当月供货数据与酒店当月应付账款金额一致",
                        "4. 供应商根据核对后确认的金额开具相关增值税发票给酒店",
                        "5. 请供应商在确认后，需加盖公章或财务专用章，扫描后邮件回传酒店做存档",
                        "6. 建议随确认函发送增值税发票号和发票金额以及发票复印件",
                        "7. 电子邮件发送至：",
                        "8. 本函请在收到后 2 个工作日内返回",
                        "9. 扫描件需清晰显示：金额、盖章、日期三要素，模糊文件视为无效"
                    ]
                    
                    for i, remark in enumerate(remarks):
                        cell = summary_sheet.cell(row=26+i, column=1, value=remark)
                        cell.font = remark_font
                        cell.alignment = remark_alignment
                        # 合并每行的A至F列，但跳过第32行（26+6）
                        if 26+i != 32:
                            summary_sheet.merge_cells(start_row=26+i, start_column=1, end_row=26+i, end_column=6)
                    
                    # 在B32单元格中添加邮箱地址
                    email_cell = summary_sheet.cell(row=32, column=2, value=email_address)
                    email_cell.font = remark_font
                    email_cell.alignment = remark_alignment
                    
                    # 在第36行A列插入供应商确认日期文字
                    date_font = Font(size=11)
                    date_alignment = Alignment(horizontal='left', vertical='center')
                    
                    date_cell = summary_sheet.cell(row=36, column=1, value="供应商确认日期：_______年_______月_______日")
                    date_cell.font = date_font
                    date_cell.alignment = date_alignment
                    # 合并供应商确认日期行的A至F列
                    summary_sheet.merge_cells(start_row=36, start_column=1, end_row=36, end_column=6)
                    # 合并第39行的A至F列
                    summary_sheet.merge_cells(start_row=39, start_column=1, end_row=39, end_column=6)
                    
                    # 在第38行插入供应商盖章确认文字
                    stamp_font = Font(size=13, underline="single")
                    stamp_alignment = Alignment(horizontal='center', vertical='center')
                    
                    stamp_cell = summary_sheet.cell(row=39, column=1, value="供应商盖章确认")
                    stamp_cell.font = stamp_font
                    stamp_cell.alignment = stamp_alignment
                    # 合并第39行的A至F列
                    summary_sheet.merge_cells(start_row=39, start_column=1, end_row=39, end_column=6)
                    
                    # 设置所有数据单元格的边框和对齐方式
                    from openpyxl.styles import Border, Side
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # 设置所有单元格的边框和格式
                    for row in range(14, row_idx + 1):
                        for col in range(1, 5):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.border = thin_border
                            
                            # 为数字列设置对齐方式和数字格式
                            if col > 1:  # 金额列
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                                cell.number_format = '#,##0.00'
                            else:  # 分类列
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 为分类行添加交替背景色
                    light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    
                    # 员工餐厅分类行
                    start_row = 5  # 员工餐厅分类开始行
                    for i, _ in enumerate(ordered_categories):
                        if i % 2 == 1:  # 偶数行添加浅色背景
                            for col in range(1, 5):
                                summary_sheet.cell(row=start_row + i, column=col).fill = light_fill
                    
                    # 其他餐厅（营业点）分类行
                    start_row = 5 + len(ordered_categories) + 3  # 其他餐厅（营业点）分类开始行
                    for i, _ in enumerate(ordered_categories):
                        if i % 2 == 1:  # 偶数行添加浅色背景
                            for col in range(1, 5):
                                summary_sheet.cell(row=start_row + i, column=col).fill = light_fill
                    
                    # 将"汇总"sheet更名为"确认函"
                    summary_sheet.title = "确认函"
                    self.log_message(f"已将汇总sheet更名为确认函")
                    
                    # 重新设置第14行和第15行居中对齐，浅蓝色背景色
                    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    for row in range(14, 16):
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.fill = light_blue_fill
                    self.log_message(f"已重新设置第14行和第15行居中对齐，浅蓝色背景色")
                    
                    # 设置第2行到第12行无背景色
                    for row in range(2, 13):
                        for col in range(1, 7):
                            cell = summary_sheet.cell(row=row, column=col)
                            cell.fill = PatternFill(fill_type=None)
                    self.log_message(f"已设置第2行到第12行无背景色")
                    
                    # 设置第2行、第5行、第8行和第13行的行高为30
                    for row_num in [2, 5, 8, 13]:
                        summary_sheet.row_dimensions[row_num].height = 30
                    self.log_message(f"已设置第2行、第5行、第8行和第13行的行高为30")
                    
                    # 保存文件
                    wb.save(output_file)
                    if self.edit_in_place_var.get():
                        self.log_message(f"已保留原始格式直接修改原文件")
                    else:
                        self.log_message(f"已保留原始格式保存文件到: {output_file}")
                    self.log_message(f"已创建供应商对账确认函sheet")
                except Exception as e:
                    self.log_message(f"保留格式保存失败，将使用标准方式保存: {str(e)}")
                    # 如果上面的方法失败，使用pandas直接保存
                    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    if self.edit_in_place_var.get():
                        self.log_message(f"已使用标准方式直接修改原文件")
                    else:
                        self.log_message(f"已使用标准方式保存文件到: {output_file}")
            except Exception as e:
                self.log_message(f"保存文件时出错: {str(e)}")
                return False
            
            if self.edit_in_place_var.get():
                self.log_message(f"分类完成，已直接修改原文件")
                self.log_message(f"文件路径: {output_file}")
            else:
                self.log_message(f"分类完成，文件已保存")
                self.log_message(f"文件路径: {output_file}")
            
            # 统计各分类数量和金额
            self.log_message("\n分类统计结果:")
            total_items = len(df)
            
            # 按财务标记分类统计，按指定顺序显示
            ordered_categories = ["干货", "海鲜", "酒类", "饮料", "水", "其他"]
            
            # 定义员工餐厅和其他餐厅（营业点）
            employee_restaurants = ["员工餐厅", "员工食堂"]
            
            # 初始化总计变量
            total_employee_untaxed = 0
            total_employee_tax = 0
            total_other_untaxed = 0
            total_other_tax = 0
            
            # 员工餐厅统计
            self.log_message("\n员工餐厅:")
            employee_items = len(df[df["部门"].isin(employee_restaurants)])
            
            for category in ordered_categories:
                # 筛选该分类的员工餐厅数据
                category_df = df[(df[classification_column] == category) & 
                                 (df["部门"].isin(employee_restaurants))]
                count = len(category_df)
                
                # 计算未税金额和税额
                untaxed_amount = category_df["小计金额(结算)"].sum() if not category_df.empty else 0
                tax_amount = category_df["税额(结算)"].sum() if not category_df.empty else 0
                total_amount = untaxed_amount + tax_amount
                
                # 更新员工餐厅总计
                total_employee_untaxed += untaxed_amount
                total_employee_tax += tax_amount
                
                # 输出统计信息
                percentage = (count / employee_items) * 100 if employee_items > 0 else 0
                self.log_message(f"{category}: {count}项 ({percentage:.1f}%)")
                self.log_message(f"  未税金额: {untaxed_amount:.2f}")
                self.log_message(f"  税额: {tax_amount:.2f}")
                self.log_message(f"  总金额: {total_amount:.2f}")
            
            # 员工餐厅小计
            self.log_message("\n员工餐厅小计:")
            self.log_message(f"未税金额: {total_employee_untaxed:.2f}")
            self.log_message(f"税额: {total_employee_tax:.2f}")
            self.log_message(f"总金额: {(total_employee_untaxed + total_employee_tax):.2f}")
            
            # 其他餐厅（营业点）统计
            self.log_message("\n其他餐厅（营业点）:")
            other_items = len(df[~df["部门"].isin(employee_restaurants)])
            
            for category in ordered_categories:
                # 筛选该分类的其他餐厅（营业点）数据
                category_df = df[(df[classification_column] == category) & 
                                 (~df["部门"].isin(employee_restaurants))]
                count = len(category_df)
                
                # 计算未税金额和税额
                untaxed_amount = category_df["小计金额(结算)"].sum() if not category_df.empty else 0
                tax_amount = category_df["税额(结算)"].sum() if not category_df.empty else 0
                total_amount = untaxed_amount + tax_amount
                
                # 更新其他餐厅（营业点）总计
                total_other_untaxed += untaxed_amount
                total_other_tax += tax_amount
                
                # 输出统计信息
                percentage = (count / other_items) * 100 if other_items > 0 else 0
                self.log_message(f"{category}: {count}项 ({percentage:.1f}%)")
                self.log_message(f"  未税金额: {untaxed_amount:.2f}")
                self.log_message(f"  税额: {tax_amount:.2f}")
                self.log_message(f"  总金额: {total_amount:.2f}")
            
            # 其他餐厅（营业点）小计
            self.log_message("\n其他餐厅（营业点）小计:")
            self.log_message(f"未税金额: {total_other_untaxed:.2f}")
            self.log_message(f"税额: {total_other_tax:.2f}")
            self.log_message(f"总金额: {(total_other_untaxed + total_other_tax):.2f}")
            
            # 输出总计信息
            total_untaxed = total_employee_untaxed + total_other_untaxed
            total_tax = total_employee_tax + total_other_tax
            
            self.log_message("\n总计:")
            self.log_message(f"未税金额: {total_untaxed:.2f}")
            self.log_message(f"税额: {total_tax:.2f}")
            self.log_message(f"总应付金额: {(total_untaxed + total_tax):.2f}")
            
            # 如果是批处理模式，直接返回成功
            if is_batch:
                return True
            # 非批处理模式下，询问用户是否打开文件夹
            message = "文件处理完成，" + ("已直接修改原文件" if self.edit_in_place_var.get() else f"已保存到:\n{output_file}")
            if messagebox.askyesno("处理完成", f"{message}\n\n是否打开文件所在文件夹？"):
                try:
                    output_dir = os.path.dirname(output_file)
                    if sys.platform == "win32":
                        os.startfile(output_dir)
                    elif sys.platform == "darwin":  # macOS
                        subprocess.call(["open", output_dir])
                    else:  # Linux
                        subprocess.call(["xdg-open", output_dir])
                except Exception as e:
                    self.log_message(f"无法打开文件夹: {str(e)}")
                    messagebox.showerror("错误", f"无法打开文件夹:\n{str(e)}")
            
            return True
            
        except Exception as e:
            self.log_message(f"处理文件时出错: {str(e)}")
            return False
        finally:
            if not is_batch:
                self.processing = False
                self.process_btn.config(state=NORMAL)
                self.progress['value'] = 100
    
    def bring_to_front(self):
        """将窗口带到前台"""
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after_idle(self.root.attributes, '-topmost', False)
    
    def create_developer_label(self):
        """在窗口底部创建开发者信息标签"""
        developer_frame = ttk.Frame(self.main_frame)
        developer_frame.pack(side=BOTTOM, fill=X, pady=5)
        
        developer_label = ttk.Label(
            developer_frame,
            text="Powered By Cayman Fu @ Sofitel HAIKOU 2025 Ver 2.0",
            font=("微软雅黑", 8),
            foreground="gray"
        )
        developer_label.pack(side=BOTTOM, pady=5)

if __name__ == "__main__":
    root = Tk()
    app = ProductClassificationApp(root)
    root.mainloop()
